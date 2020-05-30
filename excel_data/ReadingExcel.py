import openpyxl
import os

def excelDataReader():
	path_file = os.path.relpath('../excel_data/testdata_check.xlsx')
	vk  = openpyxl.load_workbook(path_file)
	sheet_ob = vk.active
	#sh = vk['Sheet1']
	r = sheet_ob.max_row
	print(r)
	li = []
	li1 = []
	for i in range(1,r+1):
		li=[]
		un = sheet_ob.cell(i,1)
		sh = sheet_ob.cell(i,2)
		li.insert(0, un.value)
		li1.insert(i-1,li)

	print(li1)
	return li1


test = excelDataReader()
#print(test)